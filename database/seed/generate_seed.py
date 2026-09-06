#!/usr/bin/env python3
"""Generate database/flyway/sql_repeatable/R__seed_dev_data.sql from Logan PL.xlsx.

The workbook is the source of truth for the seeded program. Each sheet tab is a
4-week training *block*, named by its Monday start date (MMDDYYYY). This script
parses every block, lays the blocks onto a single global week timeline (with the
two rest/deload weeks that make the calendar dates line up), and emits the full
program tree (blocks -> weeks -> days -> exercises -> set groups -> sets) plus
the logged sessions/set_logs for the weeks that have already happened.

Re-run after editing the spreadsheet:

    python database/seed/generate_seed.py

Everything downstream (models, API, frontend) is unaffected; only the seed SQL
is rewritten. Static rows (owner user + exercise catalog) are carried over
verbatim from the current seed file so the exercise catalog is never touched.

Idempotency: every row uses a deterministic uuid5 of a stable key and every
insert ends in `on conflict (id) do nothing`, matching the repeatable-migration
contract. For a destructive refresh use `docker compose down -v` then db-up.
"""

from __future__ import annotations

import datetime as dt
import re
import uuid
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────
HERE = Path(__file__).resolve().parent
XLSX = HERE / "Logan PL.xlsx"
SEED = HERE.parent / "flyway" / "sql_repeatable" / "R__seed_dev_data.sql"

# ── Fixed ids carried across regenerations ───────────────────────────────────
PROGRAM_ID = "8d645f69-e0a2-4b07-a30b-0a20634e2abb"
USER_ID = "265f6d7d-c361-4189-ac41-3f053b2b217d"
NS = uuid.UUID(PROGRAM_ID)  # uuid5 namespace so ids are stable & program-scoped

PROGRAM_NAME = "Logan PL — 2026"
PROGRAM_DESC = (
    "5-block powerlifting macrocycle (22 weeks) parsed from Logan PL.xlsx. "
    "Blocks 1–2 run 4 training days/week, blocks 3–5 run 3. Rest/deload weeks "
    "at week 9 (6/29) and week 14 (8/3 vacation) make the block start dates line up."
)
START_DATE = dt.date(2026, 5, 4)  # week 1, day 1 (Monday)

LB_PER_KG = 2.20462  # matches frontend/src/lib/units.ts


def uid(*parts) -> str:
    return str(uuid.uuid5(NS, ":".join(str(p) for p in parts)))


def kg(lb: float | None):
    return None if lb is None else round(lb / LB_PER_KG, 2)


# ── Column geometry ──────────────────────────────────────────────────────────
# Each week is a 6-column group. Base col (1-indexed) of the "Sets" column:
WEEK_BASE = {1: 4, 2: 11, 3: 18, 4: 25}  # D, K, R, Y
# offsets from base: 0=Sets 1=Reps 2=Intensity/Weight 3=Load Cap 4=Load Used 5=Last Set RPE


# Corrections for garbage characters typed into the source spreadsheet. Keyed
# by (sheet, row, column letter) -> intended value. Each is confirmed by the
# surrounding cells (and matched the original hand-built seed). Kept explicit
# and auditable rather than guessed at parse time.
OVERRIDES = {
    ("612026", 10, "F"): 5,  # RPE 5 (Last Set RPE in col I is 5); sheet had '�'
    ("612026", 12, "E"): 3,  # 1×3 deadlift opener; sheet had '`'
}


def cell(ws, row, base, off):
    letter = get_column_letter(base + off)
    key = (ws.title, row, letter)
    if key in OVERRIDES:
        return OVERRIDES[key]
    return ws.cell(row=row, column=base + off).value


# ── Block / week timeline ────────────────────────────────────────────────────
# sheet, block sequence, and which day-section labels in the sheet are the real
# training days (with their calendar day index Mon=0 within the 7-day week).
BLOCKS = [
    # sheet,     bseq, training sections -> calendar day index (Mon=0)
    ("542026", 1, [("Day 1", 0), ("Day 2", 2), ("Day 3", 3), ("Day 4", 4)]),
    ("612026", 2, [("Day 1", 0), ("Day 2", 2), ("Day 3", 3), ("Day 4", 4)]),
    ("762026", 3, [("Day 1", 0), ("Day 3", 2), ("Day 5", 4)]),
    ("832026", 4, [("Day 1", 0), ("Day 3", 2), ("Day 5", 4)]),
    ("972026", 5, [("Day 1", 0), ("Day 3", 2), ("Day 5", 4)]),
]

# Global week timeline. Each entry: (week_no, block_seq, kind, sheet, sheet_week)
# kind: "prog" = a programmed week (sheet_week 1..4), "deload" = full rest week.
WEEKS: list[tuple[int, int, str, str | None, int | None]] = []


def _build_timeline():
    w = 0

    def add_prog(sheet, bseq, count=4):
        nonlocal w
        for sw in range(1, count + 1):
            w_local = w + sw
            WEEKS.append((w_local, bseq, "prog", sheet, sw))
        return

    # B1: wk 1-4
    for sw in range(1, 5):
        WEEKS.append((sw, 1, "prog", "542026", sw))
    # B2: wk 5-8 prog + wk 9 deload
    for i, sw in enumerate(range(1, 5)):
        WEEKS.append((5 + i, 2, "prog", "612026", sw))
    WEEKS.append((9, 2, "deload", None, None))
    # B3: wk 10-13
    for i, sw in enumerate(range(1, 5)):
        WEEKS.append((10 + i, 3, "prog", "762026", sw))
    # B4: wk 14 deload (vacation) + wk 15-18 prog
    WEEKS.append((14, 4, "deload", None, None))
    for i, sw in enumerate(range(1, 5)):
        WEEKS.append((15 + i, 4, "prog", "832026", sw))
    # B5: wk 19-22
    for i, sw in enumerate(range(1, 5)):
        WEEKS.append((19 + i, 5, "prog", "972026", sw))


_build_timeline()

# Block metadata for the program_blocks table.
BLOCK_META = {
    1: ("Block 1", dt.date(2026, 5, 4)),
    2: ("Block 2", dt.date(2026, 6, 1)),
    3: ("Block 3", dt.date(2026, 7, 6)),
    4: ("Block 4", dt.date(2026, 8, 3)),
    5: ("Block 5", dt.date(2026, 9, 7)),
}

# Weeks 1-18 are history (before Block 5's 9/7 start); log sessions for them.
LAST_HISTORY_WEEK = 18

# ── Exercise catalog: sheet name -> (exercise_id, display name) ───────────────
EX = {
    "Comp Squat": ("704b039d-895c-476b-80ed-991010629bb2", "Comp Squat"),
    "Comp Bench": ("eacd9689-6804-4bb6-96db-c50a46157746", "Comp Bench"),
    "Comp Deadlift": ("2f1678ce-4c33-467b-8f5b-2d3fe99900dd", "Comp Deadlift"),
    "2ct Paused Bench": ("5b99f804-ee24-4442-b9b1-7c1059024b2c", "2ct Paused Bench"),
    "HB Squat": ("630119e6-40ba-4f1f-a8f0-2eb13a560f2e", "HB Squat"),
    "RDL": ("0bd72d61-3e3a-4b90-8eda-772e52bdb98e", "RDL"),
    "DB RDL": ("e0561fa4-9685-4544-b9bb-289426eea2fd", "DB RDL"),
    "Pec Dec": ("476a4b77-164d-44b1-b7bf-82d916315bf7", "Pec Dec"),
    "V Bar Pulldown": ("e1b0ba29-5c2d-4596-8f44-3b7c5929eecc", "V Bar Pulldown"),
    "Machine Press": ("bb3c0781-210a-4386-9f8d-19a75e1f1bc1", "Machine Press"),
    "Kelso Shrug": ("cdb350d7-da7e-4b8a-8de2-94749b23d79d", "Kelso Shrug"),
    "Cable Curl": ("c1e30306-ecc4-4a9d-83bf-c0a0cfef24d7", "Cable Curl"),
    "Tricep of choice": ("bcc50be4-55ba-4337-b92e-2cbd74aea345", "Tricep of choice"),
    "Alternating SL Quad Ext": ("f1dadafd-14bf-4b7b-9abf-fecc2aa8560f", "Alternating SL Quad Ext"),
    "Alternating SL Hamstring Curl": ("3ff72cb2-c689-404c-9948-7cd8b5a8cfef", "Alternating SL Hamstring Curl"),
}


# ── Cell parsing ─────────────────────────────────────────────────────────────
def parse_reps(v):
    """Return (min, max). Excel turns '6-10' into a June-10 datetime, so a
    datetime means month=low, day=high. '3+' (AMRAP) -> (3, None)."""
    if v is None or v == "":
        return None, None
    if isinstance(v, dt.datetime):
        return v.month, v.day
    if isinstance(v, (int, float)):
        return int(v), int(v)
    s = str(v).strip()
    if "+" in s:
        m = re.search(r"\d+", s)
        return (int(m.group()), None) if m else (None, None)
    m = re.match(r"(\d+)\s*-\s*(\d+)", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"\d+", s)
    return (int(m.group()), int(m.group())) if m else (None, None)


def _fmt_num(x: float) -> str:
    return str(int(x)) if float(x).is_integer() else str(x)


def _lead_lb(s) -> float | None:
    m = re.search(r"-?\d+(?:\.\d+)?", str(s))
    return float(m.group()) if m else None


def classify_intensity(f):
    """Return (intensity_text, rpe, weight_lb, kind).

    kind in {rpe, pct, weight, text, none}. Distinguishes the Intensity/Weight
    column's several meanings: a plain number ≤10 is an RPE target, a negative
    number is a back-off percentage, a '###lb' string is an absolute weight,
    and RIR/RPE strings are free-text targets."""
    if f is None or f == "":
        return None, None, None, "none"
    if isinstance(f, (int, float)):
        x = float(f)
        if x < 0:
            pct = round(-x * 100, 1)
            return f"-{_fmt_num(pct)}%", None, None, "pct"
        if 0 < x <= 10:
            return _fmt_num(x), x, None, "rpe"
        return f"{_fmt_num(x)}lb", None, x, "weight"  # bare number >10 = weight
    s = str(f).strip()
    up = s.upper()
    if "RIR" in up or "RPE" in up:
        return s, None, None, "text"
    if re.match(r"-?\d", s):  # leading number -> weight string
        return s, None, _lead_lb(s), "weight"
    return s, None, None, "text"


def _lb_weight(v):
    """Absolute load in lb from a prescription cell, but ONLY when the cell
    literally names a weight ('lb'). Bare numbers in the Intensity/Weight and
    Load Cap columns are references (a back-off's starting point, a percent
    annotation like '315 (-5%)', or a stray figure), not a prescribed absolute
    load — the coach writes '290lb' when a load is actually prescribed. (Actual
    logged loads in the Load Used column are numeric and handled separately.)"""
    if isinstance(v, str) and "lb" in v.lower():
        return _lead_lb(v)
    return None


def resolve_load(kind, f_raw, g_val):
    """Return (prescribed_load_lb, cap_load_lb) from the intensity kind, any
    weight in Intensity/Weight, and the Load Cap column."""
    g_lb = _lb_weight(g_val)
    if kind == "weight":
        return _lb_weight(f_raw), g_lb
    if kind == "rpe":
        return g_lb, g_lb  # RPE set: programmed load = the load cap
    if kind == "pct":
        return None, None  # back-off relative to the top set
    return g_lb, g_lb  # text / none: use the cap if one is given


# ── Parse one exercise block within a day section ────────────────────────────
class Group:
    __slots__ = ("rows",)

    def __init__(self):
        self.rows = []  # each: dict per sheet row (a set group)


def parse_day_section(ws, r0, r1):
    """Parse rows (r0+1 .. r1-1) into a list of exercises, each a list of
    set-group rows. Each row carries per-week cells."""
    exercises = []  # list of (sheet_name, ex_id, display, [rows])
    cur = None
    for r in range(r0 + 1, r1):
        b = ws.cell(row=r, column=2).value
        bname = b.strip() if isinstance(b, str) else None
        # Does this row have any Sets value across the four weeks?
        has_sets = any(
            isinstance(cell(ws, r, WEEK_BASE[w], 0), (int, float))
            for w in (1, 2, 3, 4)
        )
        if bname in EX:
            # A contiguous run of rows sharing the same exercise name is ONE
            # program_exercise with multiple set groups. The sheet repeats the
            # name on every set-group row (e.g. two "Comp Squat" rows = a top
            # set + a back-off), so only start a new exercise when the name
            # actually changes from the current block.
            if cur is None or bname != cur["sheet_name"]:
                ex_id, disp = EX[bname]
                cur = {"sheet_name": bname, "ex_id": ex_id, "disp": disp, "rows": []}
                exercises.append(cur)
        elif bname and not has_sets:
            continue  # a note line (e.g. "-Work on setup..."), not an exercise
        if not has_sets or cur is None:
            continue
        rest_c = ws.cell(row=r, column=3).value  # Rest Time(mins)
        rest_s = int(rest_c * 60) if isinstance(rest_c, (int, float)) and rest_c else None
        row = {"rest_s": rest_s, "weeks": {}}
        for w in (1, 2, 3, 4):
            base = WEEK_BASE[w]
            sets = cell(ws, r, base, 0)
            if not isinstance(sets, (int, float)):
                continue
            reps_min, reps_max = parse_reps(cell(ws, r, base, 1))
            f = cell(ws, r, base, 2)
            g = cell(ws, r, base, 3)
            itext, rpe, f_lb, kind = classify_intensity(f)
            load_lb, cap_lb = resolve_load(kind, f, g)
            row["weeks"][w] = {
                "sets": int(sets),
                "reps_min": reps_min,
                "reps_max": reps_max,
                "itext": itext,
                "rpe": rpe,
                "load_lb": load_lb,
                "cap_lb": cap_lb,
                "used_lb": _lead_lb(cell(ws, r, base, 4)) if cell(ws, r, base, 4) not in (None, "") else None,
                "last_rpe": cell(ws, r, base, 5) if isinstance(cell(ws, r, base, 5), (int, float)) else None,
            }
        cur["rows"].append(row)
    return exercises


def day_sections(ws):
    """Map training-section label -> (start_row, end_row)."""
    labels = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and (a.strip().startswith("Day") or a.strip() == "OFF"):
            labels.append((r, a.strip()))
    labels.append((ws.max_row + 1, "END"))
    out = {}
    for i in range(len(labels) - 1):
        r0, lab = labels[i]
        r1, _ = labels[i + 1]
        out[lab] = (r0, r1)
    return out


# ── SQL emission helpers ─────────────────────────────────────────────────────
def sv(x):
    if x is None:
        return "null"
    if isinstance(x, bool):
        return "true" if x else "false"
    if isinstance(x, (int, float)):
        return _fmt_num(x) if isinstance(x, float) else str(x)
    return "'" + str(x).replace("'", "''") + "'"


def table(cols, rows, name):
    if not rows:
        return f"-- (no {name} rows)\n"
    head = f"insert into {name}\n  ({', '.join(cols)})\nvalues\n"
    body = ",\n".join("  (" + ", ".join(sv(c) for c in row) + ")" for row in rows)
    return head + body + "\non conflict (id) do nothing;\n"


def dt_at(week_no, cal_idx, hh, mm):
    d = START_DATE + dt.timedelta(days=(week_no - 1) * 7 + cal_idx)
    return f"{d.isoformat()} {hh:02d}:{mm:02d}:00+00"


# ── Build everything ─────────────────────────────────────────────────────────
def build():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    parsed = {}  # sheet -> {section_label: [exercises]}
    for sheet, _, sections in BLOCKS:
        ws = wb[sheet]
        secs = day_sections(ws)
        parsed[sheet] = {
            lab: parse_day_section(ws, *secs[lab]) for lab, _ in sections
        }

    blocks_rows = []
    weeks_rows = []
    days_rows = []
    pex_rows = []
    grp_rows = []
    pset_rows = []
    sess_rows = []
    sex_rows = []
    setlog_rows = []

    # program_blocks
    block_weeks = {b: [w[0] for w in WEEKS if w[1] == b] for b in BLOCK_META}
    for bseq, (bname, _) in BLOCK_META.items():
        blocks_rows.append((uid("block", bseq), PROGRAM_ID, bseq, bname))

    training_by_block = {b: secs for (_, b, secs) in BLOCKS}
    sheet_by_block = {b: s for (s, b, _) in BLOCKS}

    for (week_no, bseq, kind, sheet, sheet_week) in WEEKS:
        week_id = uid("week", week_no)
        wname = f"Week {week_no}"
        if kind == "deload":
            wname = "Deload"
        weeks_rows.append((week_id, PROGRAM_ID, uid("block", bseq), week_no, wname))

        if kind == "deload":
            # 7 rest days, no exercises.
            for dseq in range(1, 8):
                off_n = "" if dseq == 1 else f" {dseq - 1}"
                days_rows.append(
                    (uid("day", week_no, dseq), week_id, dseq, "Rest", f"OFF{off_n}", True, None)
                )
            continue

        training = training_by_block[bseq]  # [(section_label, cal_idx), ...]
        # calendar idx -> (training day number, section label)
        idx_to_day = {cal: (i + 1, lab) for i, (lab, cal) in enumerate(training)}

        off_counter = 0
        for dseq in range(1, 8):
            cal_idx = dseq - 1
            day_id = uid("day", week_no, dseq)
            if cal_idx not in idx_to_day:
                off_counter += 1
                off_n = "" if off_counter == 1 else f" {off_counter}"
                days_rows.append((day_id, week_id, dseq, "Rest", f"OFF{off_n}", True, None))
                continue
            day_no, section_label = idx_to_day[cal_idx]
            day_name = f"Day {day_no}"
            days_rows.append((day_id, week_id, dseq, day_name, day_name, False, None))

            exercises = parsed[sheet][section_label]
            ex_seq = 0
            # session bookkeeping
            sess_id = uid("session", week_no, dseq)
            se_rows_local = []
            sl_rows_local = []
            any_logged = False

            for ex in exercises:
                # groups present for this sheet_week
                wk_groups = [row["weeks"][sheet_week] for row in ex["rows"] if sheet_week in row["weeks"]]
                if not wk_groups:
                    continue
                ex_seq += 1
                pex_id = uid("pex", week_no, dseq, ex_seq)
                rest_s = next((row["rest_s"] for row in ex["rows"] if row["rest_s"]), None)
                pex_rows.append((pex_id, day_id, ex_seq, ex["ex_id"], None, rest_s))

                sex_id = uid("sex", week_no, dseq, ex_seq)
                se_rows_local.append(
                    (sex_id, sess_id, ex_seq, ex["ex_id"], ex["disp"], None, None)
                )

                grp_seq = 0
                sl_seq = 0  # set_logs.sequence runs continuously across an
                #             exercise's groups (unique per session_exercise),
                #             unlike program_sets.sequence which restarts per group.
                for row in ex["rows"]:
                    if sheet_week not in row["weeks"]:
                        continue
                    cellw = row["weeks"][sheet_week]
                    grp_seq += 1
                    grp_id = uid("grp", week_no, dseq, ex_seq, grp_seq)
                    grp_rows.append((grp_id, pex_id, grp_seq))
                    n_sets = cellw["sets"]
                    load_kg = kg(cellw["load_lb"])
                    cap_kg = kg(cellw["cap_lb"])
                    used_kg = kg(cellw["used_lb"])
                    for s in range(1, n_sets + 1):
                        pset_id = uid("pset", week_no, dseq, ex_seq, grp_seq, s)
                        pset_rows.append(
                            (pset_id, grp_id, s, "working",
                             cellw["reps_min"], cellw["reps_max"], cellw["itext"],
                             load_kg, cap_kg, cellw["rpe"])
                        )
                        # set_log (history weeks only)
                        if week_no <= LAST_HISTORY_WEEK:
                            sl_seq += 1
                            logged = used_kg is not None
                            any_logged = any_logged or logged
                            reps_actual = cellw["reps_max"] or cellw["reps_min"]
                            # The sheet records one RPE per set group (its "Last
                            # Set RPE"); apply it to every logged set in the group.
                            actual_rpe = cellw["last_rpe"] if logged else None
                            state = "completed" if logged else "skipped"
                            completed_at = dt_at(week_no, cal_idx, 17, 36) if logged else None
                            sl_rows_local.append(
                                (uid("setlog", week_no, dseq, ex_seq, grp_seq, s),
                                 sex_id, USER_ID, ex["ex_id"], sl_seq, "working",
                                 cellw["reps_min"], cellw["reps_max"],
                                 load_kg, cellw["rpe"], cellw["itext"],
                                 reps_actual if logged else None,
                                 used_kg if logged else None,
                                 actual_rpe,
                                 completed_at, state)
                            )

            # Emit a session only for history days that had at least one logged set.
            if week_no <= LAST_HISTORY_WEEK and any_logged:
                sess_rows.append(
                    (sess_id, USER_ID, day_id, PROGRAM_NAME, f"Day {day_no}", "completed",
                     dt_at(week_no, cal_idx, 17, 30), dt_at(week_no, cal_idx, 18, 45), None)
                )
                sex_rows.extend(se_rows_local)
                setlog_rows.extend(sl_rows_local)

    return dict(
        blocks=blocks_rows, weeks=weeks_rows, days=days_rows, pex=pex_rows,
        grp=grp_rows, pset=pset_rows, sess=sess_rows, sex=sex_rows, setlog=setlog_rows,
    )


# ── Static preamble (owner user + exercise catalog) carried over verbatim ─────
def static_preamble() -> str:
    text = SEED.read_text(encoding="utf-8")
    start = text.index("-- 1. Owner user")
    end = text.index("-- 3. Program")
    return text[start:end].rstrip() + "\n"


HEADER = """\
-- Repeatable migration: seed data for local development.
--
-- Generated from database/seed/generate_seed.py against Logan PL.xlsx.
-- Re-run that script after editing the spreadsheet to refresh this file.
--
-- Idempotency model:
--   Every row uses a deterministic UUID derived from uuid5() of a stable key,
--   and every insert ends in `on conflict (id) do nothing`. Re-running this
--   migration on a populated DB is a no-op; missing rows are filled in.
--   For a destructive refresh (e.g. after restructuring the spreadsheet),
--   wipe the data volume with `docker compose down -v` so Flyway's
--   `clean migrate` reapplies everything from scratch.
--
-- Flyway wraps each migration in its own transaction; no explicit begin/commit needed.

set search_path to fitlytics, public;

"""


def main():
    d = build()
    out = [HEADER, static_preamble(), "\n"]

    out.append("-- 3. Program\n")
    out.append(table(
        ["id", "owner_user_id", "name", "description", "start_date"],
        [(PROGRAM_ID, USER_ID, PROGRAM_NAME, PROGRAM_DESC, START_DATE.isoformat())],
        "programs",
    ))
    out.append("\n-- 3b. Training blocks (mesocycles grouping the weeks)\n")
    out.append(table(["id", "program_id", "sequence", "name"], d["blocks"], "program_blocks"))
    out.append("\n-- 4. Weeks (global sequence; deloads at wk 9 & 14)\n")
    out.append(table(
        ["id", "program_id", "program_block_id", "sequence", "name"], d["weeks"], "program_weeks"
    ))
    out.append("\n-- 5. Days (7 per week; rest days flagged)\n")
    out.append(table(
        ["id", "program_week_id", "sequence", "name", "tag", "is_rest_day", "notes"],
        d["days"], "program_days",
    ))
    out.append("\n-- 6. Program exercises\n")
    out.append(table(
        ["id", "program_day_id", "sequence", "exercise_id", "sub_text", "rest_seconds"],
        d["pex"], "program_exercises",
    ))
    out.append("\n-- 7. Program set groups\n")
    out.append(table(["id", "program_exercise_id", "sequence"], d["grp"], "program_set_groups"))
    out.append("\n-- 7b. Program sets\n")
    out.append(table(
        ["id", "group_id", "sequence", "set_type", "reps_min", "reps_max",
         "intensity_text", "prescribed_load_kg", "cap_load_kg", "prescribed_rpe"],
        d["pset"], "program_sets",
    ))
    out.append("\n-- 8. Logged sessions for weeks 1-18 (history). Load Used / Last Set RPE\n")
    out.append("--    from the sheet; sets with no logged load are marked skipped.\n")
    out.append(table(
        ["id", "user_id", "program_day_id", "program_name_snapshot", "day_name_snapshot",
         "state", "started_at", "completed_at", "notes"],
        d["sess"], "sessions",
    ))
    out.append("\n")
    out.append(table(
        ["id", "session_id", "sequence", "exercise_id", "exercise_name_snapshot",
         "sub_snapshot", "rest_seconds_snapshot"],
        d["sex"], "session_exercises",
    ))
    out.append("\n")
    out.append(table(
        ["id", "session_exercise_id", "user_id", "exercise_id", "sequence", "set_type",
         "reps_target_min", "reps_target_max", "prescribed_load_kg", "prescribed_rpe",
         "intensity_text", "reps_actual", "actual_load_kg", "actual_rpe",
         "completed_at", "state"],
        d["setlog"], "set_logs",
    ))

    with open(SEED, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("".join(out))

    # Console summary
    print(f"wrote {SEED}")
    for k in ("blocks", "weeks", "days", "pex", "grp", "pset", "sess", "sex", "setlog"):
        print(f"  {k:8s}: {len(d[k])}")


if __name__ == "__main__":
    main()
